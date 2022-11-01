import asyncio
import datetime
import pandas as pd
import openpyxl


main_data = {}

from bs4 import BeautifulSoup
import aiohttp

wb = openpyxl.load_workbook('output.xlsx')
sheet = wb.active
row_counter = 1
sheet.cell(int(row_counter), 1,'Date')
sheet.cell(int(row_counter), 2,'Event Data')
sheet.cell(int(row_counter), 3,'Event Type')

def write_excel(date,event,event_type):
    global row_counter
    row_counter +=1
    sheet.cell(int(row_counter), 1, date)
    sheet.cell(int(row_counter), 2, event)
    sheet.cell(int(row_counter), 3, event_type)


async def process_link(url,date,event):
    async with aiohttp.ClientSession() as session:
        async with session.get(url=url) as response:
            if response.status == 200:
                await parse_html(await response.text(),url,date,event)
            else:
                print(date,'############################Error',response.status)


async def parse_html(html,url,date,event):
    soup = BeautifulSoup(html, features="html.parser")
    events = soup.find_all('li', {'class': 'person'})
    if len(events) <3 or event == None:
        events = soup.find_all('li', {'class': 'event'})
        if len(events) <3 or event == None:
            write_excel(date,'NO DATA',event)
    event_data  = 'N/A'
    year = 'N/A'
    for el in events:
        try:
            data = str(el.text).split(' ')
            year = data[0]
            data.pop(0)
            event_data = ' '.join(el for el in data)
        except:
            pass
        # try:
        #     year = el.find('a').text()
        # except:
        #     year = 'N/A'
        # try:
        #     event_data = str(el.text).replace(str(year),'')
        # except:
        #     pass
        write_excel(date+year,str(event_data),event)
    pass

async def main():
    #iterate all days
    start_date = datetime.date(2020, 1, 1)
    end_date = datetime.date(2020, 12, 31)
    delta = datetime.timedelta(days=1)
    while (start_date <= end_date):
        month = start_date.strftime('%B %d %Y').split(' ')[0]
        date_without_year = f'{month}-{start_date.day}-'
        start_date += delta
        events = ['birthdays','deaths','events']
        for event in events:
            url = f'https://www.onthisday.com/history/{event}/{month}/{start_date.day}'
            await process_link(url,date_without_year,event)
            print(month,start_date.day)
            print(event)
        print('##################################')
    pass

event_loop = asyncio.get_event_loop()
event_loop.run_until_complete(main())
# df = pd.DataFrame(main_data)
wb.save('output.xlsx')
# df.to_excel('output.xlsx')

#url that has no data acts like this: https://www.onthisday.com/history/weddings-divorces/december/13


#GOOD URL FOR PARSING PURPOSES: https://www.onthisday.com/history/deaths/december/12